import csv
import json
import os
import logging
from datetime import datetime
from typing import Any

logger = logging.getLogger(__name__)

def export_research(data: list[dict[str, Any]], output_path: str, format: str = "auto") -> str:
    """Export research data to a file. Format auto-detected from extension or specified."""
    if format == "auto":
        ext = os.path.splitext(output_path)[1].lower()
        format = {"xlsx": "excel", ".xlsx": "excel", ".csv": "csv", ".json": "json", ".md": "markdown"}.get(ext, "excel")

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    if format == "excel":
        return _export_excel(data, output_path)
    elif format == "csv":
        return _export_csv(data, output_path)
    elif format == "json":
        return _export_json(data, output_path)
    elif format == "markdown":
        return _export_markdown(data, output_path)
    else:
        raise ValueError(f"Unsupported format: {format}")

def _export_excel(data: list[dict], path: str) -> str:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        logger.warning("openpyxl not installed, falling back to CSV")
        return _export_csv(data, path.replace(".xlsx", ".csv"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Research Results"

    if not data:
        wb.save(path)
        return path

    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Write headers
    headers = list(data[0].keys())
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header.replace("_", " ").title())
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Write data
    alt_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    for row_idx, row_data in enumerate(data, 2):
        for col, header in enumerate(headers, 1):
            value = row_data.get(header, "")
            if isinstance(value, (list, dict)):
                value = json.dumps(value, ensure_ascii=False)
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # Auto-width columns (approximate)
    for col, header in enumerate(headers, 1):
        max_len = len(header)
        for row in data:
            val = str(row.get(header, ""))
            max_len = max(max_len, min(len(val), 60))
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = max_len + 4

    # Freeze header row
    ws.freeze_panes = "A2"

    # Add auto-filter
    ws.auto_filter.ref = ws.dimensions

    wb.save(path)
    logger.info(f"Exported {len(data)} rows to {path}")
    return path

def _export_csv(data: list[dict], path: str) -> str:
    if not data:
        with open(path, "w") as f:
            pass
        return path
    headers = list(data[0].keys())
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        for row in data:
            # Serialize complex values
            clean = {}
            for k, v in row.items():
                clean[k] = json.dumps(v, ensure_ascii=False) if isinstance(v, (list, dict)) else v
            writer.writerow(clean)
    logger.info(f"Exported {len(data)} rows to {path}")
    return path

def _export_json(data: list[dict], path: str) -> str:
    with open(path, "w", encoding="utf-8") as f:
        json.dump({"exported_at": datetime.utcnow().isoformat(), "count": len(data), "results": data}, f, indent=2, ensure_ascii=False, default=str)
    logger.info(f"Exported {len(data)} rows to {path}")
    return path

def _export_markdown(data: list[dict], path: str) -> str:
    if not data:
        with open(path, "w") as f:
            f.write("# Research Results\n\nNo data.\n")
        return path
    headers = list(data[0].keys())
    lines = [f"# Research Results\n", f"*Exported: {datetime.utcnow().isoformat()}*\n"]
    # Table header
    lines.append("| " + " | ".join(h.replace("_", " ").title() for h in headers) + " |")
    lines.append("| " + " | ".join("---" for _ in headers) + " |")
    for row in data:
        vals = []
        for h in headers:
            v = row.get(h, "")
            if isinstance(v, (list, dict)):
                v = json.dumps(v, ensure_ascii=False)
            vals.append(str(v).replace("|", "\\|").replace("\n", " "))
        lines.append("| " + " | ".join(vals) + " |")
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")
    logger.info(f"Exported {len(data)} rows to {path}")
    return path
